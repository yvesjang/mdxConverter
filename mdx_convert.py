"""
mdx_convert.py — Markdown <-> Excel 双向转换工具

用法:
  python mdx_convert.py to-xl  input.md       # 每张表 → 独立 .xlsx 文件
  python mdx_convert.py to-md  input.xlsx     # 每个 Sheet → 独立 .md 文件
"""

import sys
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────
#  工具函数
# ──────────────────────────────────────────────

def _strip_md_markup(text: str) -> str:
    """去除 Markdown / Obsidian 行内标记，保留纯文本。"""
    text = re.sub(r'==(.+?)==', r'\1', text)            # Obsidian 高亮
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)        # 粗体 **
    text = re.sub(r'__(.+?)__', r'\1', text)             # 粗体 __
    text = re.sub(r'\*(.+?)\*', r'\1', text)             # 斜体 *
    text = re.sub(r'(?<!\w)_(.+?)_(?!\w)', r'\1', text) # 斜体 _（保留 snake_case）
    text = re.sub(r'`(.+?)`', r'\1', text)               # 行内代码
    text = re.sub(r'\[(.+?)\]\(.+?\)', r'\1', text)      # 链接
    return text.strip()


def _safe_filename(name: str, max_len: int = 60) -> str:
    """将任意字符串转为合法文件名。"""
    name = re.sub(r'[\\/*?:"<>|]', '_', name)
    name = name.strip('. ')
    return name[:max_len] or 'table'


def _is_table_row(line: str) -> bool:
    return line.strip().startswith('|')


def _is_separator_row(line: str) -> bool:
    stripped = line.strip().strip('|')
    return bool(re.match(r'^[\-\:\|\s]+$', stripped)) and '-' in stripped


def _split_row(line: str) -> list[str]:
    parts = line.strip().strip('|').split('|')
    return [_strip_md_markup(p) for p in parts]


def _collect_table_lines(lines: list[str], start: int) -> tuple[list[str], int]:
    """
    从 start 行开始收集连续表格行。
    遇到空行时向前看：若空行后面仍有表格行则继续收集（跳过空行）。
    """
    table_lines = []
    i = start
    n = len(lines)

    while i < n:
        line = lines[i]
        if _is_table_row(line):
            table_lines.append(line)
            i += 1
            continue
        if line.strip() == '':
            j = i + 1
            while j < n and lines[j].strip() == '':
                j += 1
            if j < n and _is_table_row(lines[j]):
                i = j   # 跳过空行，继续
                continue
        break

    return table_lines, i


def _parse_table_block(lines: list[str]) -> pd.DataFrame | None:
    """将一段表格行解析为 DataFrame。"""
    if not lines:
        return None

    sep_idx = next((i for i, l in enumerate(lines) if _is_separator_row(l)), None)
    if sep_idx is None or sep_idx == 0:
        return None

    headers = _split_row(lines[sep_idx - 1])
    ncols = len(headers)
    header_raw = set(p.strip() for p in lines[sep_idx - 1].strip('|').split('|'))

    data_rows = []
    for line in lines[sep_idx + 1:]:
        if _is_separator_row(line):
            continue
        if set(p.strip() for p in line.strip('|').split('|')) == header_raw:
            continue
        row = _split_row(line)
        if len(row) < ncols:
            row += [''] * (ncols - len(row))
        data_rows.append(row[:ncols])

    if not data_rows:
        return None

    return pd.DataFrame(data_rows, columns=headers)


def _try_cast(df: pd.DataFrame) -> pd.DataFrame:
    """尝试将纯数字列转为数值类型。"""
    for col in df.columns:
        try:
            df[col] = pd.to_numeric(df[col].replace('', None), errors='raise')
        except (ValueError, TypeError):
            pass
    return df


def _write_xlsx(df: pd.DataFrame, out_path: Path, sheet_name: str = 'Sheet1'):
    """将 DataFrame 写入格式化的 xlsx 文件。"""
    sheet_name = re.sub(r'[\\/*?\[\]:]', '_', sheet_name)[:31]
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        ws = writer.book.active
        ws.freeze_panes = 'A2'

        header_fill = PatternFill('solid', start_color='D9E1F2')
        header_font = Font(bold=True, name='Arial')
        center = Alignment(horizontal='center', vertical='center')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


# ──────────────────────────────────────────────
#  Markdown → xlsx（每张表 → 独立文件）
# ──────────────────────────────────────────────

def parse_md_tables(md_text: str) -> list[tuple[str | None, pd.DataFrame]]:
    """
    提取 Markdown 中所有表格。
    返回 [(heading_or_None, DataFrame), ...]
    heading 取紧邻表格前最近的标题行文本。
    """
    results = []
    last_heading = None
    lines = md_text.splitlines()
    i = 0

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        m = re.match(r'^#{1,6}\s+(.*)', stripped)
        if m:
            last_heading = m.group(1).strip()
            i += 1
            continue

        if _is_table_row(line):
            table_lines, i = _collect_table_lines(lines, i)
            df = _parse_table_block(table_lines)
            if df is not None:
                results.append((last_heading, df))
            continue

        i += 1

    return results


def md_to_xlsx(md_path: str):
    md_path = Path(md_path)
    if not md_path.exists():
        sys.exit(f"[错误] 文件不存在: {md_path}")

    text = md_path.read_text(encoding='utf-8')
    tables = parse_md_tables(text)

    if not tables:
        sys.exit("[错误] 未在 Markdown 文件中找到任何表格")

    stem = md_path.stem
    out_dir = md_path.parent
    used_names: set[str] = set()

    for idx, (heading, df) in enumerate(tables, start=1):
        df = _try_cast(df)

        # 决定输出文件名
        if len(tables) == 1:
            # 只有一张表：直接用原文件名
            base_name = stem
        elif heading:
            base_name = f"{stem}_{_safe_filename(heading)}"
        else:
            base_name = f"{stem}_table_{idx}"

        # 避免重名
        file_name = base_name
        counter = 2
        while file_name in used_names:
            file_name = f"{base_name}_{counter}"
            counter += 1
        used_names.add(file_name)

        out_path = out_dir / f"{file_name}.xlsx"
        sheet_name = _safe_filename(heading) if heading else 'Sheet1'
        _write_xlsx(df, out_path, sheet_name=sheet_name)
        print(f"[完成] {out_path}  ({len(df)} 行)")

    if len(tables) > 1:
        print(f"\n共提取 {len(tables)} 张表格 → {len(tables)} 个 xlsx 文件")


# ──────────────────────────────────────────────
#  xlsx → Markdown（每个 Sheet → 独立文件）
# ──────────────────────────────────────────────

def _df_to_md_table(df: pd.DataFrame) -> str:
    """DataFrame → Markdown 表格字符串。用 iloc 按位置取列，兼容重复列名。"""
    cols = list(df.columns)
    ncols = len(cols)

    # 按列位置计算最大宽度，避免重复列名导致 df[c] 返回 DataFrame
    col_widths = []
    for i in range(ncols):
        col_vals = df.iloc[:, i].astype(str).tolist()
        col_widths.append(max(max((len(v) for v in col_vals), default=0), len(str(cols[i])), 3))

    def fmt_row(values):
        return '| ' + ' | '.join(str(v).ljust(col_widths[i]) for i, v in enumerate(values)) + ' |'

    sep = '| ' + ' | '.join('-' * w for w in col_widths) + ' |'
    data_rows = [fmt_row(df.iloc[r].tolist()) for r in range(len(df))]
    return '\n'.join([fmt_row(cols), sep] + data_rows)


def xlsx_to_md(xlsx_path: str):
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        sys.exit(f"[错误] 文件不存在: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    stem = xlsx_path.stem
    out_dir = xlsx_path.parent

    written = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        rows = [
            ['' if v is None else str(v) for v in row]
            for row in ws.iter_rows(values_only=True)
            if any(v is not None for v in row)
        ]
        if not rows:
            print(f"[跳过] Sheet '{sheet_name}' 无数据")
            continue

        df = pd.DataFrame(rows[1:], columns=rows[0])

        # 文件命名：单 Sheet 用原文件名，多 Sheet 加 Sheet 名
        if len(wb.sheetnames) == 1:
            out_name = f"{stem}.md"
        else:
            out_name = f"{stem}_{_safe_filename(sheet_name)}.md"

        out_path = out_dir / out_name
        content = f"## {sheet_name}\n\n{_df_to_md_table(df)}\n"
        out_path.write_text(content, encoding='utf-8')
        written.append(out_path)
        print(f"[完成] {out_path}  ({len(df)} 行)")

    if len(written) > 1:
        print(f"\n共转换 {len(written)} 个 Sheet → {len(written)} 个 md 文件")


# ──────────────────────────────────────────────
#  入口
# ──────────────────────────────────────────────

def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)

    mode, input_path = sys.argv[1], sys.argv[2]
    if mode == 'to-xl':
        md_to_xlsx(input_path)
    elif mode == 'to-md':
        xlsx_to_md(input_path)
    else:
        sys.exit(f"[错误] 未知模式: {mode}")


if __name__ == '__main__':
    main()
